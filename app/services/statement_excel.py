"""
services/statement_excel.py — Account statement Excel + PDF generation.

Save path:
  <STATEMENT_BASE_PATH>/<YEAR>/<MM>/<COMPANY>_<YYYY-MM>.xlsx / .pdf

Folders are created automatically.
"""

import logging
from datetime import date as _date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from app.config import settings

logger = logging.getLogger(__name__)

_MONTH_NAMES = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]

_HEADERS = ["Date", "Invoice No", "LPO", "Debit (AED)", "Credit (AED)", "Balance (AED)"]

# ── Our company letterhead constants ─────────────────────────────────────────
_OUR_NAME  = "DAR AL SALAM ENG. TURNING WORKS W/SHOP L.L.C"
_OUR_TEL   = "06-5350865 / 06-5464327"
_OUR_FAX   = "06-5351865"
_OUR_POBOX = "34356"
_OUR_AREA  = "Ind. Area No. 10"
_OUR_CITY  = "Sharjah - U.A.E"
_OUR_EMAIL = "darselam@eim.ae"


def _lpo_from_desc(desc: str) -> str:
    """Return LPO number if desc is 'LPO XXXXX', otherwise '-'."""
    s = (desc or "").strip()
    if s.upper().startswith("LPO ") and len(s) > 4:
        return s[4:].strip()
    return "-"


# ── Amount-to-words (English, AED/Fils) ───────────────────────────────────────

_ONES = [
    "", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
    "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
    "Seventeen", "Eighteen", "Nineteen",
]
_TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty",
         "Sixty", "Seventy", "Eighty", "Ninety"]


def _say(n: int) -> str:
    if n == 0:
        return ""
    if n < 20:
        return _ONES[n]
    if n < 100:
        o = _ONES[n % 10]
        return _TENS[n // 10] + (" " + o if o else "")
    rest = _say(n % 100)
    return _ONES[n // 100] + " Hundred" + (" " + rest if rest else "")


def _amount_to_words(amount: float) -> str:
    """Convert a float amount to English words for AED currency."""
    dirhams = int(amount)
    fils    = round((amount - dirhams) * 100)
    if dirhams == 0 and fils == 0:
        return "Zero Dirhams Only"
    parts: list[str] = []
    if dirhams // 1_000_000:
        parts.append(_say(dirhams // 1_000_000) + " Million")
    if (dirhams % 1_000_000) // 1_000:
        parts.append(_say((dirhams % 1_000_000) // 1_000) + " Thousand")
    if dirhams % 1_000:
        parts.append(_say(dirhams % 1_000))
    dirham_words = " ".join(parts) if parts else "Zero"
    result = dirham_words + " Dirhams"
    if fils:
        result += " and " + _say(fils) + " Fils"
    return result + " Only"


# ── Path helpers ──────────────────────────────────────────────────────────────

def statement_folder(month: int, year: int) -> Path:
    mm = str(month).zfill(2)
    return Path(settings.STATEMENT_BASE_PATH) / str(year) / mm


def statement_path(company_name: str, month: int, year: int) -> Path:
    mm = str(month).zfill(2)
    safe = company_name.replace("/", "-").replace("\\", "-").replace(":", "-")
    return statement_folder(month, year) / f"{safe}_{year}-{mm}.xlsx"


# ── Workbook styling helpers ──────────────────────────────────────────────────

def _thin():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


_FILL_DARK  = PatternFill("solid", fgColor="1F4E79")
_FILL_MID   = PatternFill("solid", fgColor="2E75B6")
_FILL_GRAY  = PatternFill("solid", fgColor="F2F2F2")
_FILL_CLOSE = PatternFill("solid", fgColor="D6E4F0")


# ── Date parsing ──────────────────────────────────────────────────────────────

def _parse_date(date_str: str) -> "datetime | None":
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(date_str).strip(), fmt)
        except (ValueError, AttributeError):
            pass
    return None


# ── Statement generation ──────────────────────────────────────────────────────

def generate_statement(
    company_name: str,
    month: int,
    year: int,
    mode: str = "full",
) -> "tuple[Path, Path | None]":
    """
    Build an account statement for company_name for the given month/year.

    mode:
      "full"   — all debit + credit entries in the period
      "unpaid" — only entries whose invoice still has an outstanding balance
      "pdf"    — same as full + attempt PDF conversion

    Returns (xlsx_path, pdf_path | None).
    Raises FileNotFoundError if no ledger exists for the company.
    """
    from app.services.ledger_excel import get_ledger_summary

    summary = get_ledger_summary(company_name)
    if not summary:
        raise FileNotFoundError(f"No ledger found for {company_name!r}")

    all_rows = summary["rows"]

    # ── Split rows into pre-period (opening) and in-period ────────────────────
    opening_balance = 0.0
    period_rows: list[dict] = []

    for row in all_rows:
        dt = _parse_date(row.get("date", ""))
        if dt is None:
            continue
        if dt.year < year or (dt.year == year and dt.month < month):
            opening_balance += row["debit"] - row["credit"]
        elif dt.year == year and dt.month == month:
            period_rows.append(row)
        # rows after the period are ignored

    # ── Unpaid filter: keep only entries for invoices with net balance > 0 ────
    if mode == "unpaid":
        inv_net: dict[str, float] = {}
        for row in period_rows:
            inv = row.get("invoice", "")
            inv_net[inv] = inv_net.get(inv, 0.0) + row["debit"] - row["credit"]
        period_rows = [r for r in period_rows if inv_net.get(r.get("invoice", ""), 0.0) > 0.01]

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement"

    # Page layout: A4 portrait, fit to 1 page wide (height flows naturally from top)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize   = 9       # A4
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0       # let height flow — no vertical shrink/stretch
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.5, bottom=0.5, header=0.2, footer=0.2
    )
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered   = False  # content starts from top, no padding

    month_label = _MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
    mode_label  = " (Unpaid Only)" if mode == "unpaid" else ""
    _al_l = Alignment(horizontal="left",   vertical="center")
    _al_c = Alignment(horizontal="center", vertical="center")
    _al_r = Alignment(horizontal="right",  vertical="center")

    # ── Rows 1–4 : Company letterhead ─────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"].value     = _OUR_NAME
    ws["A1"].font      = Font(bold=True, size=15, color="1F4E79")
    ws["A1"].alignment = _al_l
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:F2")
    ws["A2"].value     = f"Tel: {_OUR_TEL}    Fax: {_OUR_FAX}"
    ws["A2"].font      = Font(size=9)
    ws["A2"].alignment = _al_l

    ws.merge_cells("A3:F3")
    ws["A3"].value     = f"P.O. Box: {_OUR_POBOX}    {_OUR_AREA}    {_OUR_CITY}"
    ws["A3"].font      = Font(size=9)
    ws["A3"].alignment = _al_l

    ws.merge_cells("A4:F4")
    ws["A4"].value     = f"E-mail: {_OUR_EMAIL}"
    ws["A4"].font      = Font(size=9)
    ws["A4"].alignment = _al_l

    # Blue separator below row 4
    _sep = Border(bottom=Side(border_style="medium", color="1F4E79"))
    for col in range(1, 7):
        ws.cell(row=4, column=col).border = _sep

    # Row 5 — spacer between letterhead and title
    ws.row_dimensions[5].height = 8

    # ── Row 6 : Title ─────────────────────────────────────────────────────────
    ws.merge_cells("A6:F6")
    ws["A6"].value     = "STATEMENT OF ACCOUNT"
    ws["A6"].font      = Font(bold=True, size=13, color="1F4E79")
    ws["A6"].alignment = _al_c
    ws.row_dimensions[6].height = 24

    # ── Row 7 : To + Period ───────────────────────────────────────────────────
    ws.merge_cells("A7:C8")
    ws["A7"].value     = f"To:\n{company_name}"
    ws["A7"].font      = Font(bold=True, size=10)
    ws["A7"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[7].height = 16
    ws.row_dimensions[8].height = 18

    ws.merge_cells("D7:F8")
    ws["D7"].value     = (
        f"Period: {month_label} {year}{mode_label}\n"
        f"Date: {_date.today().strftime('%d-%m-%Y')}"
    )
    ws["D7"].font      = Font(size=10)
    ws["D7"].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

    # ── Row 9 : Opening Balance ───────────────────────────────────────────────
    _OB = 9
    ws.merge_cells(f"A{_OB}:D{_OB}")
    ws.cell(row=_OB, column=1, value="Opening Balance:")
    ws.cell(row=_OB, column=1).font      = Font(bold=True, size=10)
    ws.cell(row=_OB, column=1).fill      = _FILL_GRAY
    ws.cell(row=_OB, column=1).alignment = _al_r
    ws.merge_cells(f"E{_OB}:F{_OB}")
    ws.cell(row=_OB, column=5, value=f"AED {opening_balance:,.2f}")
    ws.cell(row=_OB, column=5).font      = Font(bold=True, size=10)
    ws.cell(row=_OB, column=5).fill      = _FILL_GRAY
    ws.cell(row=_OB, column=5).alignment = _al_r
    for col in range(1, 7):
        ws.cell(row=_OB, column=col).border = _thin()
    ws.row_dimensions[_OB].height = 18

    # ── Row 10 : Column headers ───────────────────────────────────────────────
    _HDR = 10
    for col, label in enumerate(_HEADERS, 1):
        c = ws.cell(row=_HDR, column=col, value=label)
        c.font      = Font(bold=True, size=10, color="FFFFFF")
        c.fill      = _FILL_MID
        c.alignment = _al_c
        c.border    = _thin()
    ws.row_dimensions[_HDR].height = 20

    # ── Rows 11+ : Data ───────────────────────────────────────────────────────
    _D0 = 11
    running = opening_balance
    for i, row in enumerate(period_rows):
        r   = _D0 + i
        d   = float(row.get("debit",  0) or 0)
        cr  = float(row.get("credit", 0) or 0)
        lpo = _lpo_from_desc(row.get("desc", ""))
        running += d - cr

        ws.cell(row=r, column=1, value=row.get("date",    "")).font = Font(size=10)
        ws.cell(row=r, column=1).alignment = _al_c
        ws.cell(row=r, column=2, value=row.get("invoice", "")).font = Font(size=10)
        ws.cell(row=r, column=2).alignment = _al_c
        ws.cell(row=r, column=3, value=lpo).font = Font(size=10)
        ws.cell(row=r, column=3).alignment = _al_c
        if d:
            ws.cell(row=r, column=4, value=round(d, 2)).font = Font(size=10)
        if cr:
            ws.cell(row=r, column=5, value=round(cr, 2)).font = Font(size=10)
        ws.cell(row=r, column=4).alignment = _al_r
        ws.cell(row=r, column=5).alignment = _al_r
        ws.cell(row=r, column=6, value=round(running, 2)).font = Font(size=10, bold=True)
        ws.cell(row=r, column=6).alignment = _al_r
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = _thin()
        ws.row_dimensions[r].height = 18

    # ── Total row ─────────────────────────────────────────────────────────────
    _tot = _D0 + len(period_rows)
    ws.merge_cells(f"A{_tot}:E{_tot}")
    ws.cell(row=_tot, column=1, value="TOTAL  (AED)")
    ws.cell(row=_tot, column=1).font      = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=_tot, column=1).fill      = _FILL_DARK
    ws.cell(row=_tot, column=1).alignment = _al_r
    ws.cell(row=_tot, column=6, value=round(running, 2))
    ws.cell(row=_tot, column=6).font      = Font(bold=True, size=11, color="FFFFFF")
    ws.cell(row=_tot, column=6).fill      = _FILL_DARK
    ws.cell(row=_tot, column=6).alignment = _al_r
    for col in range(1, 7):
        ws.cell(row=_tot, column=col).border = _thin()
    ws.row_dimensions[_tot].height = 20

    # ── Amount in words ───────────────────────────────────────────────────────
    _words = _tot + 1
    ws.merge_cells(f"A{_words}:F{_words}")
    ws.cell(row=_words, column=1,
            value=f"Amount in Words:  AED {_amount_to_words(abs(running))}")
    ws.cell(row=_words, column=1).font      = Font(italic=True, size=10)
    ws.cell(row=_words, column=1).alignment = _al_l
    ws.row_dimensions[_words].height = 16

    # ── Signature block ───────────────────────────────────────────────────────
    _sig = _words + 2
    ws.merge_cells(f"D{_sig}:F{_sig}")
    ws.cell(row=_sig, column=4, value="for DAR AL SALAM ENG. TURNING WORKS W/SHOP")
    ws.cell(row=_sig, column=4).font      = Font(bold=True, size=9)
    ws.cell(row=_sig, column=4).alignment = _al_c
    # underline via bottom border on signature row
    for col in range(4, 7):
        ws.cell(row=_sig, column=col).border = Border(
            bottom=Side(border_style="thin", color="000000")
        )
    ws.merge_cells(f"D{_sig + 1}:F{_sig + 1}")
    ws.cell(row=_sig + 1, column=4, value="Authorised Signatory")
    ws.cell(row=_sig + 1, column=4).font      = Font(size=8, color="555555")
    ws.cell(row=_sig + 1, column=4).alignment = _al_c

    # ── Column widths ─────────────────────────────────────────────────────────
    # Proportional for A4 portrait: Date(12%) Invoice(15%) LPO(18%) Debit(18%) Credit(18%) Balance(19%)
    # Total ~100 units to span full page width at 0.4in margins
    for col, width in enumerate([12, 15, 18, 18, 18, 19], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Print area — exact content range, no trailing empty rows ─────────────
    ws.print_area = f"A1:F{_sig + 1}"

    # ── Save Excel ────────────────────────────────────────────────────────────
    out_path = statement_path(company_name, month, year)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info("Statement saved: %s", out_path)

    # ── PDF conversion via LibreOffice ────────────────────────────────────────
    pdf_path: "Path | None" = None
    from app.services.pdf_export import export_to_pdf
    pdf_result = export_to_pdf(out_path)
    if pdf_result["status"] == "created" and pdf_result["pdf_path"]:
        pdf_path = Path(pdf_result["pdf_path"])
        logger.info("Statement PDF: %s", pdf_path)

    return out_path, pdf_path


# ── Summary text (used by bot reply) ─────────────────────────────────────────

def statement_summary_text(
    company_name: str,
    month: int,
    year: int,
    opening: float,
    closing: float,
    row_count: int,
) -> str:
    month_label = _MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
    outstanding = round(closing, 2)
    return (
        f"📄 <b>Account Statement</b>\n\n"
        f"Company:  <b>{company_name}</b>\n"
        f"Period:   <b>{month_label} {year}</b>\n"
        f"Entries:  {row_count}\n"
        f"Opening:  AED {opening:,.2f}\n"
        f"Closing:  <b>AED {closing:,.2f}</b>\n"
        f"{'⚠️ Outstanding: ' if outstanding > 0 else '✅ Outstanding: '}"
        f"<b>AED {outstanding:,.2f}</b>"
    )
