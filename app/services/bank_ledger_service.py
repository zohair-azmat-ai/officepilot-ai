"""
services/bank_ledger_service.py — Bank account ledger: Excel storage, balance, statements.

Storage layout:
  BANK_LEDGER_BASE_PATH/
    BANK_LEDGER_2026.xlsx    ← one workbook per year, sheet "Transactions"
    2026/04/                 ← generated statement files
      DD-MM-YYYY BANK STATEMENT APRIL 2026.xlsx
      DD-MM-YYYY BANK STATEMENT APRIL 2026.pdf

Sheet columns:
  A  Txn ID        B  Date        C  Type         D  Mode
  E  Party/Company F  Description G  In           H  Out
  I  Balance       J  Notes
"""

import logging
from datetime import date as _date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)

# ── Column indices (1-based, for openpyxl) ────────────────────────────────────
_C_TXN   = 1   # A  — transaction ID (TXN-0001 …)
_C_DATE  = 2   # B
_C_TYPE  = 3   # C
_C_MODE  = 4   # D
_C_PARTY = 5   # E
_C_DESC  = 6   # F
_C_IN    = 7   # G
_C_OUT   = 8   # H
_C_BAL   = 9   # I
_C_NOTES = 10  # J

_HEADERS = ["Txn ID", "Date", "Type", "Mode", "Party/Company",
            "Description", "In", "Out", "Balance", "Notes"]

_MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"]

# ── Styles ────────────────────────────────────────────────────────────────────
_HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
_IN_FILL   = PatternFill("solid", fgColor="E8F5E9")   # light green
_OUT_FILL  = PatternFill("solid", fgColor="FFEBEE")   # light red
_THIN      = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")
_MONEY  = '#,##0.00'

_COL_WIDTHS = [10, 14, 12, 18, 25, 30, 14, 14, 16, 20]   # A-J


# ── Helpers ───────────────────────────────────────────────────────────────────

def _ledger_path(year: int) -> Path:
    base = Path(settings.BANK_LEDGER_BASE_PATH)
    base.mkdir(parents=True, exist_ok=True)
    return base / f"BANK_LEDGER_{year}.xlsx"


def _style_header_row(ws, row: int = 1) -> None:
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row, col)
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN


def _style_data_row(ws, row: int, is_incoming: bool) -> None:
    fill = _IN_FILL if is_incoming else _OUT_FILL
    for col in range(1, len(_HEADERS) + 1):
        cell = ws.cell(row, col)
        cell.fill   = fill
        cell.border = _THIN
        if col in (_C_IN, _C_OUT, _C_BAL):
            cell.number_format = _MONEY
            cell.alignment     = _RIGHT
        elif col in (_C_TXN, _C_DATE):
            cell.alignment = _CENTER
        else:
            cell.alignment = _LEFT


def _open_or_create(year: int):
    path = _ledger_path(year)
    if path.exists():
        wb = openpyxl.load_workbook(path)
        ws = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active
        return wb, ws, path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"
    for col_i, header in enumerate(_HEADERS, 1):
        ws.cell(1, col_i).value = header
    _style_header_row(ws, 1)
    ws.row_dimensions[1].height = 20
    for i, w in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    wb.save(path)
    return wb, ws, path


def _last_balance(ws) -> float:
    last = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[_C_BAL - 1] is not None:
            try:
                last = float(row[_C_BAL - 1])
            except (TypeError, ValueError):
                pass
    return last


def _next_row(ws) -> int:
    """Return the row number for the next entry (one after the last non-empty data row)."""
    last = 1  # row 1 is always the header
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if any(v is not None for v in row):
            last = i
    return last + 1


def _last_txn_number(ws) -> int:
    """Return the highest TXN sequence number already in the sheet (0 if none)."""
    last = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[_C_TXN - 1]
        if val and isinstance(val, str) and val.upper().startswith("TXN-"):
            try:
                n = int(val[4:])
                if n > last:
                    last = n
            except ValueError:
                pass
    return last


def _recalculate_balances(ws) -> None:
    """Recompute the running Balance column for every non-empty data row in sheet order."""
    balance = 0.0
    for row_cells in ws.iter_rows(min_row=2):
        if not any(c.value for c in row_cells):
            continue
        try:
            in_val  = float(row_cells[_C_IN  - 1].value or 0)
        except (TypeError, ValueError):
            in_val = 0.0
        try:
            out_val = float(row_cells[_C_OUT - 1].value or 0)
        except (TypeError, ValueError):
            out_val = 0.0
        balance = round(balance + in_val - out_val, 2)
        row_cells[_C_BAL - 1].value = balance


def _parse_date(s: str) -> "_date | None":
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(s).strip(), fmt).date()
        except ValueError:
            pass
    return None


# ── Public API ────────────────────────────────────────────────────────────────

def add_bank_entry(
    transaction_type: str,
    mode: str,
    party: str = "",
    description: str = "",
    amount_in: float = 0.0,
    amount_out: float = 0.0,
    notes: str = "",
    date_str: "str | None" = None,
) -> float:
    """
    Append a row to the correct year's bank ledger and return the new balance.
    If date_str is provided it determines which year's workbook receives the entry.
    All text fields are stored uppercase for consistency.
    """
    today = _date.today()
    if date_str:
        parsed_d = _parse_date(date_str)
        year = parsed_d.year if parsed_d else today.year
    else:
        year     = today.year
        date_str = today.strftime("%d-%m-%Y")

    wb, ws, path = _open_or_create(year)
    prev_bal = _last_balance(ws)
    new_bal  = round(prev_bal + amount_in - amount_out, 2)

    txn_no = _last_txn_number(ws) + 1
    txn_id = f"TXN-{txn_no:04d}"

    row_num = _next_row(ws)
    ws.cell(row_num, _C_TXN).value   = txn_id
    ws.cell(row_num, _C_DATE).value  = date_str
    ws.cell(row_num, _C_TYPE).value  = transaction_type
    ws.cell(row_num, _C_MODE).value  = mode
    ws.cell(row_num, _C_PARTY).value = party.upper()       if party       else ""
    ws.cell(row_num, _C_DESC).value  = description.upper() if description else ""
    ws.cell(row_num, _C_IN).value    = amount_in  if amount_in  else None
    ws.cell(row_num, _C_OUT).value   = amount_out if amount_out else None
    ws.cell(row_num, _C_BAL).value   = new_bal
    ws.cell(row_num, _C_NOTES).value = notes.upper() if notes else ""
    # Safety: clear any stray values beyond the 10-column layout
    ws.cell(row_num, 11).value = None
    ws.cell(row_num, 12).value = None
    _style_data_row(ws, row_num, transaction_type == "Incoming")

    wb.save(path)
    logger.info(
        "Bank entry: %s %s %s  party=%r  in=%.2f  out=%.2f  bal=%.2f  row=%d",
        txn_id, transaction_type, mode, party, amount_in, amount_out, new_bal, row_num,
    )
    return new_bal, txn_id


def get_bank_balance(year: "int | None" = None) -> float:
    """Return the running balance from the last row of the ledger."""
    year = year or _date.today().year
    path = _ledger_path(year)
    if not path.exists():
        return 0.0
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active
    return _last_balance(ws)


def check_duplicate(party: str, amount: float, date_str: str) -> bool:
    """
    Return True if a row with the same date, party (case-insensitive), and
    amount (±0.01) already exists in the ledger for that date's year.
    """
    today   = _date.today()
    parsed  = _parse_date(date_str) if date_str else None
    year    = parsed.year if parsed else today.year
    eff_dt  = date_str or today.strftime("%d-%m-%Y")

    path = _ledger_path(year)
    if not path.exists():
        return False

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        r_date   = str(row[_C_DATE  - 1] or "").strip()
        r_party  = str(row[_C_PARTY - 1] or "").strip().upper()
        r_in     = float(row[_C_IN  - 1] or 0)
        r_out    = float(row[_C_OUT - 1] or 0)
        r_amount = r_in or r_out
        if (r_date == eff_dt and
                r_party == party.strip().upper() and
                abs(r_amount - amount) < 0.01):
            return True
    return False


def delete_bank_entry(txn_id: str, year: "int | None" = None) -> "dict | None":
    """
    Remove the row matching txn_id and recalculate all running balances.
    Returns a dict summarising the deleted row, or None if not found.
    """
    year = year or _date.today().year
    wb, ws, path = _open_or_create(year)

    target_row   = None
    deleted_info: dict = {}
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[_C_TXN - 1] or "").upper() == txn_id.upper():
            target_row = i
            deleted_info = {
                "txn_id": str(row[_C_TXN   - 1] or ""),
                "date":   str(row[_C_DATE  - 1] or ""),
                "type":   str(row[_C_TYPE  - 1] or ""),
                "party":  str(row[_C_PARTY - 1] or ""),
                "in":     float(row[_C_IN  - 1] or 0),
                "out":    float(row[_C_OUT - 1] or 0),
            }
            break

    if target_row is None:
        return None

    ws.delete_rows(target_row)
    _recalculate_balances(ws)
    wb.save(path)
    logger.info("Deleted bank entry: %s", txn_id)
    return deleted_info


def edit_bank_entry(
    txn_id: str,
    field:  str,
    value:  str,
    year:   "int | None" = None,
) -> "dict | None":
    """
    Update one field of an existing entry and recalculate all running balances.

    Supported fields: amount | in | out | party | notes | description | date
    Returns {"txn_id", "field", "value"} on success, None if not found / bad field.
    """
    year = year or _date.today().year
    wb, ws, path = _open_or_create(year)

    target_row = None
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[_C_TXN - 1] or "").upper() == txn_id.upper():
            target_row = i
            break

    if target_row is None:
        return None

    f = field.lower()

    if f == "amount":
        try:
            amt = float(value.replace(",", ""))
        except ValueError:
            return None
        txn_type = str(ws.cell(target_row, _C_TYPE).value or "").lower()
        if "incoming" in txn_type:
            ws.cell(target_row, _C_IN).value  = amt
            ws.cell(target_row, _C_OUT).value = None
        else:
            ws.cell(target_row, _C_OUT).value = amt
            ws.cell(target_row, _C_IN).value  = None

    elif f == "in":
        try:
            ws.cell(target_row, _C_IN).value = float(value.replace(",", ""))
        except ValueError:
            return None

    elif f == "out":
        try:
            ws.cell(target_row, _C_OUT).value = float(value.replace(",", ""))
        except ValueError:
            return None

    elif f == "party":
        ws.cell(target_row, _C_PARTY).value = value.strip().upper()

    elif f in ("notes", "note"):
        ws.cell(target_row, _C_NOTES).value = value.strip().upper()

    elif f in ("desc", "description"):
        ws.cell(target_row, _C_DESC).value = value.strip().upper()

    elif f == "date":
        ws.cell(target_row, _C_DATE).value = value.strip()

    else:
        return None

    _recalculate_balances(ws)
    wb.save(path)
    logger.info("Edited bank entry: %s  %s=%s", txn_id, field, value)
    return {"txn_id": txn_id, "field": field, "value": value}


def generate_bank_statement(
    month: int,
    year: int,
    as_pdf: bool = False,
) -> dict:
    """
    Generate a print-ready A4 portrait statement Excel (and optionally PDF).

    Fixed row layout (never uses ws.append so row numbers are deterministic):
      Row 1  : Title bar
      Row 2  : Spacer
      Rows 3-6 : Summary (Opening Balance / Total In / Total Out / Closing Balance)
      Row 7  : Spacer
      Row 8  : Table header
      Row 9+ : Transaction rows

    Print settings: A4 portrait, fit-to-1-page-wide, fit-to-1-page-tall when
    ≤ 25 data rows (forces single-page PDF for small/test data).

    Returns dict with keys:
      xlsx_path, pdf_path, opening_balance, total_in, total_out,
      closing_balance, row_count, month_name
    """
    from app.services.pdf_export import export_to_pdf
    from openpyxl.worksheet.properties import PageSetupProperties

    month_name  = _MONTH_NAMES[month] if 1 <= month <= 12 else str(month)
    ledger_path = _ledger_path(year)

    # ── Read all rows from this year's ledger ─────────────────────────────────
    all_rows: list[dict] = []
    if ledger_path.exists():
        wb_src = openpyxl.load_workbook(ledger_path, data_only=True)
        ws_src = wb_src["Transactions"] if "Transactions" in wb_src.sheetnames else wb_src.active
        for row in ws_src.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            all_rows.append({
                "txn_id":      str(row[_C_TXN   - 1] or ""),
                "date":        str(row[_C_DATE  - 1] or "").strip(),
                "type":        str(row[_C_TYPE  - 1] or ""),
                "mode":        str(row[_C_MODE  - 1] or ""),
                "party":       str(row[_C_PARTY - 1] or ""),
                "description": str(row[_C_DESC  - 1] or ""),
                "in":          float(row[_C_IN   - 1] or 0),
                "out":         float(row[_C_OUT  - 1] or 0),
                "balance":     float(row[_C_BAL  - 1] or 0),
                "notes":       str(row[_C_NOTES - 1] or ""),
            })

    # Opening balance = balance of last row before this month
    opening_balance = 0.0
    for r in all_rows:
        d = _parse_date(r["date"])
        if d and (d.year < year or (d.year == year and d.month < month)):
            opening_balance = r["balance"]

    period_rows = [
        r for r in all_rows
        if (d := _parse_date(r["date"])) and d.year == year and d.month == month
    ]

    total_in  = round(sum(r["in"]  for r in period_rows), 2)
    total_out = round(sum(r["out"] for r in period_rows), 2)
    closing   = round(opening_balance + total_in - total_out, 2)

    # ── Output path ───────────────────────────────────────────────────────────
    stmt_folder = Path(settings.BANK_LEDGER_BASE_PATH) / str(year) / str(month).zfill(2)
    stmt_folder.mkdir(parents=True, exist_ok=True)
    today_str = _date.today().strftime("%d-%m-%Y")
    xlsx_path = stmt_folder / f"{today_str} BANK STATEMENT {month_name.upper()} {year}.xlsx"

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name[:3]} {year}"

    # ── A4 portrait print setup ───────────────────────────────────────────────
    ws.page_setup.paperSize   = 9            # A4
    ws.page_setup.orientation = "portrait"
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    ws.print_options.horizontalCentered = True

    # Fit to page: always 1 page wide; 1 page tall for small data
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 1 if len(period_rows) <= 25 else 0

    # ── Column widths (A4 portrait proportions) ───────────────────────────────
    # Total ≈ 129 units; fitToWidth=1 scales automatically if wider than page
    _STMT_WIDTHS = [9, 12, 10, 14, 16, 20, 11, 11, 13, 12]   # A-J
    for i, w in enumerate(_STMT_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    tc = ws.cell(1, 1)
    tc.value     = f"BANK STATEMENT  {month_name.upper()} {year}"
    tc.font      = Font(bold=True, size=16, color="FFFFFF")
    tc.fill      = _HDR_FILL
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # ── Row 2: Spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[2].height = 6

    # ── Rows 3-6: Summary block ───────────────────────────────────────────────
    _SUMM_FILL    = PatternFill("solid", fgColor="EBF3FB")
    _CLOSING_FILL = PatternFill("solid", fgColor="D0E8F5")
    _SUMM_THIN    = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    summary_items = [
        ("Opening Balance:", f"AED {opening_balance:,.2f}", False),
        ("Total In:",        f"AED {total_in:,.2f}",        False),
        ("Total Out:",       f"AED {total_out:,.2f}",       False),
        ("Closing Balance:", f"AED {closing:,.2f}",         True),
    ]
    for r_idx, (label, val, is_closing) in enumerate(summary_items, 3):
        fill = _CLOSING_FILL if is_closing else _SUMM_FILL
        ws.merge_cells(f"A{r_idx}:E{r_idx}")
        ws.merge_cells(f"F{r_idx}:J{r_idx}")

        lc = ws.cell(r_idx, 1)
        lc.value     = label
        lc.font      = Font(bold=True, size=11)
        lc.fill      = fill
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border    = _SUMM_THIN

        vc = ws.cell(r_idx, 6)
        vc.value     = val
        vc.font      = Font(bold=is_closing, size=11,
                            color="1F4E79" if is_closing else "000000")
        vc.fill      = fill
        vc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        vc.border    = _SUMM_THIN
        ws.row_dimensions[r_idx].height = 22

    # ── Row 7: Spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[7].height = 6

    # ── Row 8: Table header ────────────────────────────────────────────────────
    _HDR_ROW = 8
    for c_idx, h in enumerate(_HEADERS, 1):
        cell = ws.cell(_HDR_ROW, c_idx)
        cell.value     = h
        cell.font      = _HDR_FONT
        cell.fill      = _HDR_FILL
        cell.alignment = _CENTER
        cell.border    = _THIN
    ws.row_dimensions[_HDR_ROW].height = 22

    # ── Rows 9+: Transaction data ──────────────────────────────────────────────
    for i, r in enumerate(period_rows):
        row_n = 9 + i
        ws.cell(row_n, _C_TXN).value   = r["txn_id"]
        ws.cell(row_n, _C_DATE).value  = r["date"]
        ws.cell(row_n, _C_TYPE).value  = r["type"]
        ws.cell(row_n, _C_MODE).value  = r["mode"]
        ws.cell(row_n, _C_PARTY).value = r["party"]
        ws.cell(row_n, _C_DESC).value  = r["description"]
        ws.cell(row_n, _C_IN).value    = r["in"]  or None
        ws.cell(row_n, _C_OUT).value   = r["out"] or None
        ws.cell(row_n, _C_BAL).value   = r["balance"] if r["balance"] else None
        ws.cell(row_n, _C_NOTES).value = r["notes"]
        # Safety: clear any stray values beyond the 10-column layout
        ws.cell(row_n, 11).value = None
        ws.cell(row_n, 12).value = None
        _style_data_row(ws, row_n, r["type"] == "Incoming")
        ws.row_dimensions[row_n].height = 18

    # ── Print area and repeating header ───────────────────────────────────────
    last_row = 8 + len(period_rows)
    ws.print_area        = f"A1:J{last_row}"
    ws.print_title_rows  = f"{_HDR_ROW}:{_HDR_ROW}"

    wb.save(xlsx_path)
    logger.info("Bank statement Excel: %s  rows=%d", xlsx_path, len(period_rows))

    # ── PDF export ────────────────────────────────────────────────────────────
    pdf_path = None
    if as_pdf:
        result = export_to_pdf(xlsx_path)
        if result["status"] == "created":
            pdf_path = result["pdf_path"]
            logger.info("Bank statement PDF: %s", pdf_path)

    return {
        "xlsx_path":       xlsx_path,
        "pdf_path":        pdf_path,
        "opening_balance": opening_balance,
        "total_in":        total_in,
        "total_out":       total_out,
        "closing_balance": closing,
        "row_count":       len(period_rows),
        "month_name":      month_name,
    }
