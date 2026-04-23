"""
services/ledger_excel.py — Per-company Excel ledger files.

Each company has exactly one file:
  <LEDGER_BASE_PATH>/<COMPANY NAME>.xlsx

Layout (created fresh per company):
  Row 1 : Company name header (merged A1:F1)
  Row 2 : blank
  Row 3 : Column headers — Date | Invoice No | Description | Debit (AED) | Credit (AED) | Balance (AED)
  Row 4+: Data rows

Quotation code never imports or calls this module.
"""

import logging
from datetime import date as _date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)

_HEADERS      = ["Date", "Invoice No", "Description", "Debit (AED)", "Credit (AED)", "Balance (AED)"]
_DATA_START   = 4   # first row available for data entries
_COL_DATE     = 1   # A
_COL_INVOICE  = 2   # B
_COL_DESC     = 3   # C
_COL_DEBIT    = 4   # D
_COL_CREDIT   = 5   # E
_COL_BALANCE  = 6   # F


# ── Path helpers ──────────────────────────────────────────────────────────────

def ledger_path(company_name: str) -> Path:
    return Path(settings.LEDGER_BASE_PATH) / f"{company_name}.xlsx"


def ledger_exists(company_name: str) -> bool:
    return ledger_path(company_name).exists()


# ── Workbook creation ─────────────────────────────────────────────────────────

def _thin():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def _make_workbook(company_name: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ledger"

    # Row 1 — company name header
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = company_name.upper()
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Row 2 — blank spacer

    # Row 3 — column headers
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    for col, label in enumerate(_HEADERS, 1):
        c = ws.cell(row=3, column=col, value=label)
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border    = _thin()
    ws.row_dimensions[3].height = 18

    # Column widths
    for col, width in enumerate([14, 14, 38, 16, 16, 16], 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    return wb


def create_ledger(company_name: str) -> Path:
    """
    Create a new ledger Excel file for company_name.
    Raises FileExistsError if it already exists.
    """
    path = ledger_path(company_name)
    if path.exists():
        raise FileExistsError(company_name)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = _make_workbook(company_name)
    wb.save(path)
    logger.info("Ledger created: %s", path)
    return path


# ── Row helpers ───────────────────────────────────────────────────────────────

def _next_data_row(ws) -> int:
    """Return the next empty row index at or after _DATA_START."""
    last = _DATA_START - 1
    for row in ws.iter_rows(min_row=_DATA_START, min_col=1, max_col=6):
        if any(cell.value is not None for cell in row):
            last = row[0].row
    return last + 1


def _write_row(ws, row: int, date_str: str, invoice: str,
               desc: str, debit: float | None, credit: float | None) -> None:
    ws.cell(row=row, column=_COL_DATE,    value=date_str)
    ws.cell(row=row, column=_COL_INVOICE, value=invoice)
    ws.cell(row=row, column=_COL_DESC,    value=desc)
    if debit is not None:
        ws.cell(row=row, column=_COL_DEBIT, value=debit)
    if credit is not None:
        ws.cell(row=row, column=_COL_CREDIT, value=credit)
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = _thin()


def _recalculate_balance(ws) -> None:
    """Recompute running balance top-to-bottom and write into _COL_BALANCE.

    Rule: balance += debit; balance -= credit (payment).
    Every data row gets its balance written, even if debit and credit are both zero.
    """
    running = 0.0
    for row in ws.iter_rows(min_row=_DATA_START, min_col=1, max_col=6):
        # Skip completely empty rows
        if all(cell.value is None for cell in row):
            continue
        debit  = row[_COL_DEBIT  - 1].value
        credit = row[_COL_CREDIT - 1].value
        if isinstance(debit,  (int, float)):
            running += debit
        if isinstance(credit, (int, float)):
            running -= credit
        balance_cell = row[_COL_BALANCE - 1]
        balance_cell.value  = round(running, 2)
        balance_cell.border = _thin()


# ── Public write operations ───────────────────────────────────────────────────

def add_debit_row(company_name: str, invoice_no: str,
                  amount: float, description: str = "",
                  date_str: str | None = None) -> None:
    """Append an invoice / debit row to the company ledger."""
    path = ledger_path(company_name)
    if not path.exists():
        raise FileNotFoundError(f"No ledger file for {company_name!r}. Create it first.")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    row = _next_data_row(ws)
    entry_date = date_str or _date.today().strftime("%d-%m-%Y")
    _write_row(ws, row, entry_date, invoice_no,
               description or "Entry", amount, None)
    _recalculate_balance(ws)
    wb.save(path)
    logger.info("Debit row added: %s / %s / %.2f  date=%s", company_name, invoice_no, amount, entry_date)


def add_credit_row(company_name: str, invoice_no: str, amount: float,
                   date_str: str | None = None) -> None:
    """Append a payment / credit row to the company ledger."""
    path = ledger_path(company_name)
    if not path.exists():
        raise FileNotFoundError(f"No ledger file for {company_name!r}. Create it first.")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    row = _next_data_row(ws)
    entry_date = date_str or _date.today().strftime("%d-%m-%Y")
    _write_row(ws, row, entry_date, invoice_no,
               "Payment received", None, amount)
    _recalculate_balance(ws)
    wb.save(path)
    logger.info("Credit row added: %s / %s / %.2f  date=%s", company_name, invoice_no, amount, entry_date)


def add_multi_credit_rows(
    company_name: str,
    invoices: list,
    amounts: list,
    date_str: str | None = None,
) -> list:
    """Append one credit row per invoice in a single workbook open/save cycle.

    Returns list of (invoice_no, amount) tuples that were written.
    Balance is recalculated once after all rows are appended.
    """
    path = ledger_path(company_name)
    if not path.exists():
        raise FileNotFoundError(f"No ledger file for {company_name!r}. Create it first.")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    entry_date = date_str or _date.today().strftime("%d-%m-%Y")
    written: list = []
    for inv_no, amount in zip(invoices, amounts):
        row = _next_data_row(ws)
        _write_row(ws, row, entry_date, str(inv_no), "Payment received", None, float(amount))
        written.append((str(inv_no), float(amount)))
    _recalculate_balance(ws)
    wb.save(path)
    logger.info("Multi-credit rows added: %s  invoices=%s  date=%s", company_name, invoices, entry_date)
    return written


# ── Public read operations ────────────────────────────────────────────────────

def get_ledger_summary(company_name: str) -> dict | None:
    """
    Return a summary dict:
      { company, total_debit, total_credit, outstanding, rows: [...] }
    Returns None if the ledger file does not exist.
    """
    path = ledger_path(company_name)
    if not path.exists():
        return None

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    total_debit = total_credit = 0.0
    rows: list[dict] = []

    for row in ws.iter_rows(min_row=_DATA_START, values_only=True):
        if len(row) < 5:
            continue
        date_val, invoice, desc, debit, credit, *_ = row
        if date_val is None and invoice is None:
            continue
        d = float(debit)  if isinstance(debit,  (int, float)) else 0.0
        c = float(credit) if isinstance(credit, (int, float)) else 0.0
        total_debit  += d
        total_credit += c
        rows.append({
            "date":    str(date_val) if date_val else "",
            "invoice": str(invoice)  if invoice  else "",
            "desc":    str(desc)     if desc     else "",
            "debit":   d,
            "credit":  c,
        })

    return {
        "company":     company_name,
        "total_debit":  round(total_debit,  2),
        "total_credit": round(total_credit, 2),
        "outstanding":  round(total_debit - total_credit, 2),
        "rows":         rows,
    }
