"""
create_template.py — One-time helper script to generate a starter Excel template.

Run this ONCE to create templates/quotation_template.xlsx with:
  - A professional layout
  - Labelled rows matching the cell map in config.py (B2:B11)
  - Placeholder text in each data cell so you know exactly where to look

After running this script, you can open the file in Excel, redesign the
layout (colors, logo, fonts, borders) however you like.
The ONLY requirement is that the data cells remain at the addresses listed
in config.py → CELL_MAP (defaults: B2, B3, B4, B6–B11).

Usage:
    python create_template.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path(__file__).parent / "templates" / "quotation_template.xlsx"


def thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def create_template() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18

    # ── Header banner (row 1) ──────────────────────────────────────────────────
    ws.merge_cells("A1:C1")
    header = ws["A1"]
    header.value        = "QUOTATION"
    header.font         = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
    header.fill         = PatternFill("solid", fgColor="1A1A2E")
    header.alignment    = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Data rows (labels in A, values in B) ──────────────────────────────────
    label_style = Font(name="Calibri", size=11, bold=True, color="333333")
    value_style = Font(name="Calibri", size=11, color="1A1A2E")
    label_fill  = PatternFill("solid", fgColor="F2F4F8")
    value_fill  = PatternFill("solid", fgColor="FFFFFF")

    rows = [
        (2,  "Date",        "<<date>>"),
        (3,  "Ref No.",     "<<ref_no>>"),
        (4,  "Client Name", "<<client_name>>"),
        (5,  "",            ""),               # spacer
        (6,  "Description", "<<description>>"),
        (7,  "Size",        "<<size>>"),
        (8,  "Quantity",    "<<quantity>>"),
        (9,  "Rate (AED)",  "<<rate>>"),
        (10, "Tax (AED)",   "<<tax>>"),
        (11, "Total (AED)", "<<total>>"),
    ]

    for row_num, label, placeholder in rows:
        ws.row_dimensions[row_num].height = 22

        a_cell = ws.cell(row=row_num, column=1, value=label)
        b_cell = ws.cell(row=row_num, column=2, value=placeholder)

        if label:
            a_cell.font      = label_style
            a_cell.fill      = label_fill
            a_cell.border    = thin_border()
            a_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            b_cell.font      = value_style
            b_cell.fill      = value_fill
            b_cell.border    = thin_border()
            b_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Highlight total row ────────────────────────────────────────────────────
    total_fill = PatternFill("solid", fgColor="E8F4FD")
    bold_font  = Font(name="Calibri", size=11, bold=True, color="1A1A2E")
    ws["A11"].fill = total_fill
    ws["B11"].fill = total_fill
    ws["A11"].font = bold_font
    ws["B11"].font = bold_font

    # ── Footer note ────────────────────────────────────────────────────────────
    ws.merge_cells("A13:C13")
    note = ws["A13"]
    note.value     = "Thank you for your business."
    note.font      = Font(name="Calibri", size=10, italic=True, color="888888")
    note.alignment = Alignment(horizontal="center")

    # ── Save ───────────────────────────────────────────────────────────────────
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Template created at: {OUTPUT_PATH.resolve()}")
    print()
    print("Cell mapping (matches config.py defaults):")
    print("  B2  = date")
    print("  B3  = ref_no")
    print("  B4  = client_name")
    print("  B6  = description")
    print("  B7  = size")
    print("  B8  = quantity")
    print("  B9  = rate")
    print("  B10 = tax")
    print("  B11 = total")
    print()
    print("Open the file in Excel, add your company logo / branding,")
    print("then keep those cell addresses intact (or update CELL_MAP in config.py).")


if __name__ == "__main__":
    create_template()
